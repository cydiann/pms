"""
Archive Service for Completed Requests

This module handles the archival of completed requests:
1. Generate Excel report with document hyperlinks
2. Download all documents from MinIO
3. Create ZIP archive
4. Save to local storage
5. Delete requests and documents from system
"""

import os
import zipfile
import logging
from datetime import timedelta
from pathlib import Path
from django.conf import settings
from django.utils import timezone
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import Request, ApprovalHistory, ProcurementDocument, RequestArchive
from .storage import get_storage

logger = logging.getLogger('pms.app')


class ArchiveService:
    """Service for archiving completed requests"""

    def __init__(self):
        self.archive_dir = Path(settings.BASE_DIR) / 'archives'
        self.archive_dir.mkdir(exist_ok=True)

    def get_completed_requests_in_period(self, period_start, period_end):
        """Get all completed requests within the specified period"""
        return Request.objects.filter(
            status='completed',
            updated_at__gte=period_start,
            updated_at__lte=period_end
        ).select_related('created_by', 'last_approver').order_by('created_at')

    def generate_excel_report(self, requests, workbook_path):
        """
        Generate Excel report with request details and document hyperlinks

        Returns: Path to generated Excel file
        """
        wb = Workbook()

        # Sheet 1: Requests
        ws_requests = wb.active
        ws_requests.title = "Requests"

        # Headers for requests sheet
        headers = [
            'Request Number', 'Item', 'Description', 'Quantity', 'Unit',
            'Category', 'Delivery Address', 'Reason', 'Created By',
            'Created At', 'Submitted At', 'Completed At', 'Status',
            'Approval Level', 'Revision Count', 'Document Links'
        ]

        # Style headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_num, header in enumerate(headers, 1):
            cell = ws_requests.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add request data
        storage = get_storage()
        for row_num, request in enumerate(requests, 2):
            ws_requests.cell(row=row_num, column=1, value=request.request_number)
            ws_requests.cell(row=row_num, column=2, value=request.item)
            ws_requests.cell(row=row_num, column=3, value=request.description)
            ws_requests.cell(row=row_num, column=4, value=float(request.quantity))
            ws_requests.cell(row=row_num, column=5, value=request.get_unit_display())
            ws_requests.cell(row=row_num, column=6, value=request.category)
            ws_requests.cell(row=row_num, column=7, value=request.delivery_address)
            ws_requests.cell(row=row_num, column=8, value=request.reason)
            ws_requests.cell(row=row_num, column=9, value=request.created_by.get_full_name())
            ws_requests.cell(row=row_num, column=10, value=request.created_at.strftime('%Y-%m-%d %H:%M'))
            ws_requests.cell(row=row_num, column=11, value=request.submitted_at.strftime('%Y-%m-%d %H:%M') if request.submitted_at else '')
            ws_requests.cell(row=row_num, column=12, value=request.updated_at.strftime('%Y-%m-%d %H:%M'))
            ws_requests.cell(row=row_num, column=13, value=request.get_status_display())
            ws_requests.cell(row=row_num, column=14, value=request.approval_level)
            ws_requests.cell(row=row_num, column=15, value=request.revision_count)

            # Add document hyperlinks
            documents = ProcurementDocument.objects.filter(
                request=request,
                status='uploaded'
            )

            if documents.exists() and storage:
                doc_links = []
                for doc in documents:
                    # Generate presigned URL (valid for 7 days for archive purposes)
                    try:
                        url = storage.get_presigned_download_url(doc.object_name, expiry_seconds=604800)
                        doc_links.append(f'{doc.file_name}: {url}')
                    except Exception as e:
                        logger.warning(f"Could not generate URL for document {doc.id}: {e}")
                        doc_links.append(f'{doc.file_name}: [URL generation failed]')

                cell = ws_requests.cell(row=row_num, column=16)
                cell.value = '\n'.join(doc_links)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                ws_requests.cell(row=row_num, column=16, value='No documents')

        # Auto-size columns
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            if col_num == 16:  # Document links column
                ws_requests.column_dimensions[column_letter].width = 60
            elif col_num in [3, 7, 8]:  # Description, delivery, reason
                ws_requests.column_dimensions[column_letter].width = 40
            else:
                ws_requests.column_dimensions[column_letter].width = 20

        # Sheet 2: Approval History (Complete Log)
        ws_history = wb.create_sheet(title="Approval History")

        history_headers = [
            'Request Number', 'Item', 'Action', 'Performed By', 'Username',
            'Hierarchy Level', 'Notes', 'Review Notes', 'Timestamp'
        ]

        for col_num, header in enumerate(history_headers, 1):
            cell = ws_history.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Get all approval history for these requests
        request_ids = [r.id for r in requests]
        approval_history = ApprovalHistory.objects.filter(
            request_id__in=request_ids
        ).select_related('request', 'user').order_by('request__request_number', 'created_at')

        for row_num, history in enumerate(approval_history, 2):
            ws_history.cell(row=row_num, column=1, value=history.request.request_number)
            ws_history.cell(row=row_num, column=2, value=history.request.item)
            ws_history.cell(row=row_num, column=3, value=history.get_action_display())
            ws_history.cell(row=row_num, column=4, value=history.user.get_full_name())
            ws_history.cell(row=row_num, column=5, value=history.user.username)
            ws_history.cell(row=row_num, column=6, value=history.level)
            ws_history.cell(row=row_num, column=7, value=history.notes or '')
            ws_history.cell(row=row_num, column=8, value=history.review_notes or '')
            ws_history.cell(row=row_num, column=9, value=history.created_at.strftime('%Y-%m-%d %H:%M:%S'))

        # Auto-size columns for history
        for col_num in range(1, len(history_headers) + 1):
            column_letter = get_column_letter(col_num)
            if col_num in [7, 8]:  # Notes columns
                ws_history.column_dimensions[column_letter].width = 40
            elif col_num == 2:  # Item column
                ws_history.column_dimensions[column_letter].width = 30
            else:
                ws_history.column_dimensions[column_letter].width = 20

        # Save workbook
        wb.save(workbook_path)
        logger.info(f"Excel report generated: {workbook_path}")

        return workbook_path

    def download_documents_from_minio(self, requests, temp_dir):
        """
        Download all documents for the given requests from MinIO

        Returns: Dictionary mapping request_number -> list of downloaded file paths
        """
        storage = get_storage()
        if not storage:
            logger.warning("MinIO storage not available, skipping document downloads")
            return {}

        downloaded_files = {}

        for request in requests:
            documents = ProcurementDocument.objects.filter(
                request=request,
                status='uploaded'
            )

            if not documents.exists():
                continue

            request_doc_dir = temp_dir / request.request_number
            request_doc_dir.mkdir(exist_ok=True)

            request_files = []

            for doc in documents:
                try:
                    # Download file from MinIO
                    local_file_path = request_doc_dir / doc.file_name

                    # Use MinIO client to download
                    storage.client.fget_object(
                        storage.bucket_name,
                        doc.object_name,
                        str(local_file_path)
                    )

                    request_files.append(str(local_file_path))
                    logger.info(f"Downloaded document: {doc.file_name} for {request.request_number}")

                except Exception as e:
                    logger.error(f"Failed to download document {doc.id}: {e}")

            if request_files:
                downloaded_files[request.request_number] = request_files

        return downloaded_files

    def create_zip_archive(self, excel_path, documents_by_request, zip_path):
        """
        Create ZIP archive containing Excel report and all documents

        Structure:
        archive_YYYY-MM-DD_to_YYYY-MM-DD.zip
        ├── requests_report.xlsx
        └── documents/
            ├── REQ-2025-ABC123/
            │   ├── quote.pdf
            │   └── receipt.jpg
            └── REQ-2025-XYZ789/
                └── invoice.pdf
        """
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Add Excel report
            zipf.write(excel_path, 'requests_report.xlsx')
            logger.info(f"Added Excel report to archive")

            # Add documents
            for request_number, file_paths in documents_by_request.items():
                for file_path in file_paths:
                    file_name = Path(file_path).name
                    arc_name = f'documents/{request_number}/{file_name}'
                    zipf.write(file_path, arc_name)

            logger.info(f"Added {sum(len(files) for files in documents_by_request.values())} documents to archive")

        return zip_path

    def delete_archived_requests(self, requests):
        """Delete requests and all related data from the database"""
        request_ids = [r.id for r in requests]

        with transaction.atomic():
            # Delete approval history
            ApprovalHistory.objects.filter(request_id__in=request_ids).delete()
            logger.info(f"Deleted approval history for {len(request_ids)} requests")

            # Delete procurement documents (metadata only, files handled separately)
            ProcurementDocument.objects.filter(request_id__in=request_ids).delete()
            logger.info(f"Deleted procurement document records")

            # Delete requests
            Request.objects.filter(id__in=request_ids).delete()
            logger.info(f"Deleted {len(request_ids)} requests from database")

    def delete_documents_from_minio(self, requests):
        """Delete all documents for archived requests from MinIO"""
        storage = get_storage()
        if not storage:
            logger.warning("MinIO storage not available, skipping document deletion")
            return

        for request in requests:
            documents = ProcurementDocument.objects.filter(
                request=request,
                status='uploaded'
            )

            for doc in documents:
                try:
                    storage.delete_object(doc.object_name)
                    logger.info(f"Deleted document from MinIO: {doc.object_name}")
                except Exception as e:
                    logger.error(f"Failed to delete document {doc.object_name} from MinIO: {e}")

    def create_archive(self, period_start=None, period_end=None):
        """
        Main method to create archive for completed requests

        Args:
            period_start: Start of period (defaults to 2 weeks ago)
            period_end: End of period (defaults to now)

        Returns:
            RequestArchive instance or None if no requests to archive
        """
        # Default to last 2 weeks if not specified
        if not period_end:
            period_end = timezone.now()
        if not period_start:
            period_start = period_end - timedelta(weeks=2)

        logger.info(f"Starting archive process for period {period_start} to {period_end}")

        # Get completed requests
        requests = list(self.get_completed_requests_in_period(period_start, period_end))

        if not requests:
            logger.info("No completed requests found in this period")
            return None

        logger.info(f"Found {len(requests)} completed requests to archive")

        # Generate archive filename
        archive_filename = f"archive_{period_start.date()}_{period_end.date()}.zip"
        zip_path = self.archive_dir / archive_filename

        # Create temp directory for downloads
        import tempfile
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)

            # Generate Excel report
            excel_path = temp_path / 'requests_report.xlsx'
            self.generate_excel_report(requests, excel_path)

            # Download documents from MinIO
            documents_by_request = self.download_documents_from_minio(requests, temp_path)

            # Create ZIP archive
            self.create_zip_archive(excel_path, documents_by_request, zip_path)

        # Get file size
        file_size = zip_path.stat().st_size

        # Create archive record
        archive = RequestArchive.objects.create(
            period_start=period_start,
            period_end=period_end,
            file_path=str(zip_path),
            file_size=file_size,
            request_count=len(requests),
            archived_request_ids=[r.id for r in requests],
            archived_request_numbers=[r.request_number for r in requests]
        )

        logger.info(f"Archive created successfully: {archive_filename} ({file_size} bytes)")

        # Delete requests and documents from system immediately after archive creation
        logger.info("Deleting archived requests and documents from system...")
        self.delete_documents_from_minio(requests)
        self.delete_archived_requests(requests)
        logger.info("Archived requests and documents deleted from system")

        return archive

    def cleanup_after_download(self, archive):
        """
        Clean up after archive download:
        1. Delete ZIP file from disk

        Note: Requests and documents are already deleted when archive was created
        """
        logger.info(f"Starting cleanup for archive {archive.id}")

        # Delete the ZIP file
        try:
            archive_path = Path(archive.file_path)
            if archive_path.exists():
                archive_path.unlink()
                logger.info(f"Deleted archive file: {archive.file_path}")
            else:
                logger.warning(f"Archive file not found: {archive.file_path}")
        except Exception as e:
            logger.error(f"Failed to delete archive file {archive.file_path}: {e}")

        logger.info(f"Cleanup completed for archive {archive.id}")
